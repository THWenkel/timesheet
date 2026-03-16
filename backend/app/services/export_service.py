# =============================================================================
# backend/app/services/export_service.py
#
# Export service — generates CSV, Excel (.xlsx), and PDF files from timesheet data.
#
# Each function accepts a list of TimesheetEntryRead records plus metadata
# (employee name, date range) and returns the file content as bytes.
#
# Format details:
#   CSV   — UTF-8 with BOM (ensures correct display in Microsoft Excel)
#   Excel — .xlsx via openpyxl with styled header row
#   PDF   — A4 table report via reportlab with title, metadata, and entry table
#
# The callers (export router) wrap these bytes in a FastAPI StreamingResponse
# with the appropriate Content-Type and Content-Disposition headers.
# =============================================================================

import csv
import io
from datetime import date

from app.schemas.timesheet import TimesheetEntryRead


def _minutes_to_display(minutes: int) -> str:
    """
    Convert integer minutes to 'hh:mm' display string.

    Args:
        minutes: Total minutes (must be >= 0).

    Returns:
        Zero-padded hours:minutes string, e.g. '08:30'.
    """
    h, m = divmod(minutes, 60)
    return f"{h:02d}:{m:02d}"


def generate_csv(
    entries: list[TimesheetEntryRead],
    employee_name: str,
    from_date: date,
    to_date: date,
) -> bytes:
    """
    Generate a CSV file from the given timesheet entries.

    The output is UTF-8 encoded with a BOM prefix so that Microsoft Excel
    opens it correctly without an import wizard.

    Columns: Date, Duration (hh:mm), Duration (minutes), Description

    A summary row with total minutes is appended at the end.

    Args:
        entries:       List of timesheet entries to export.
        employee_name: Display name of the employee (used in header comment).
        from_date:     Start of the export date range.
        to_date:       End of the export date range.

    Returns:
        CSV file content as bytes (UTF-8 with BOM).
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)

    # --- Metadata header rows ---
    writer.writerow(["Timesheet Export"])
    writer.writerow(["Employee:", employee_name])
    writer.writerow(["Period:", f"{from_date} - {to_date}"])
    writer.writerow([])  # blank line separator

    # --- Data header ---
    writer.writerow(["Date", "Duration (hh:mm)", "Description"])

    # --- Data rows ---
    total_minutes = 0
    for entry in entries:
        writer.writerow([
            entry.entry_date.isoformat(),
            entry.hours_display,
            entry.description,
        ])
        total_minutes += entry.minutes

    # --- Totals row ---
    writer.writerow([])
    writer.writerow([
        "TOTAL",
        _minutes_to_display(total_minutes),
        "",
    ])

    # Return UTF-8 with BOM for Excel compatibility
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def generate_excel(
    entries: list[TimesheetEntryRead],
    employee_name: str,
    from_date: date,
    to_date: date,
) -> bytes:
    """
    Generate an Excel (.xlsx) workbook from the given timesheet entries.

    Uses openpyxl to create a styled spreadsheet with:
      - A title and metadata block at the top
      - A bold header row with column widths auto-fitted
      - Data rows with alternating row colours (light blue / white)
      - A totals row at the bottom with bold formatting

    Args:
        entries:       List of timesheet entries to export.
        employee_name: Display name of the employee.
        from_date:     Start of the export date range.
        to_date:       End of the export date range.

    Returns:
        Excel file content as bytes (.xlsx format).
    """
    # Import here to avoid loading openpyxl on every request — only needed for Excel export
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"  # type: ignore[union-attr]

    # --- Colour constants ---
    header_fill = PatternFill(start_color="1B4F8A", end_color="1B4F8A", fill_type="solid")
    alt_row_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    bold_white = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    # --- Title block ---
    ws.append(["Timesheet Export"])  # type: ignore[union-attr]
    ws["A1"].font = Font(bold=True, size=14)  # type: ignore[union-attr]
    ws.append(["Employee:", employee_name])  # type: ignore[union-attr]
    ws.append(["Period:", f"{from_date} - {to_date}"])  # type: ignore[union-attr]
    ws.append([])  # type: ignore[union-attr]  blank row

    # --- Column headers (row 5) ---
    header_row = ["Date", "Duration (hh:mm)", "Description"]
    ws.append(header_row)  # type: ignore[union-attr]
    header_row_idx = ws.max_row  # type: ignore[union-attr]
    for col_idx, _ in enumerate(header_row, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx)  # type: ignore[union-attr]
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # --- Data rows ---
    total_minutes = 0
    for i, entry in enumerate(entries):
        row_data = [
            entry.entry_date.isoformat(),
            entry.hours_display,
            entry.description,
        ]
        ws.append(row_data)  # type: ignore[union-attr]
        total_minutes += entry.minutes
        # Alternate row shading for readability
        if i % 2 == 0:
            for col_idx in range(1, len(row_data) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = alt_row_fill  # type: ignore[union-attr]

    # --- Totals row ---
    ws.append([])  # type: ignore[union-attr]  blank row
    totals_row = ["TOTAL", _minutes_to_display(total_minutes), ""]
    ws.append(totals_row)  # type: ignore[union-attr]
    for col_idx in range(1, len(totals_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)  # type: ignore[union-attr]
        cell.font = bold
        cell.fill = total_fill

    # --- Column widths ---
    column_widths = [14, 18, 60]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width  # type: ignore[union-attr]

    # --- Serialize to bytes ---
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# =============================================================================
# PDF helpers — digital signature field support
# =============================================================================

def _make_sig_capture_flowable(
    Flowable: type,  # noqa: N803  — passed in after reportlab is imported
    label: str,
    height: float,
    label_style: object,
    out: list[tuple[int, float, float, float, float]],
) -> object:
    """
    Dynamically create and return a reportlab Flowable subclass instance that:
      1. Draws the label text at the bottom of the allocated cell area.
      2. Records its exact absolute page position via canvas.absolutePosition()
         so that a /Sig AcroForm widget can be added afterwards via pypdf.

    The class is defined here (not at module level) so that the reportlab
    Flowable base class is available at definition time — following the same
    deferred-import pattern used throughout this module.
    """
    from reportlab.platypus import Paragraph  # noqa: PLC0415

    class _SigCapture(Flowable):  # type: ignore[valid-type, misc]
        def __init__(self) -> None:
            super().__init__()  # type: ignore[misc]
            self.width: float = 0.0
            self.height: float = height

        def wrap(self, avail_width: float, avail_height: float) -> tuple[float, float]:  # noqa: ARG002
            self.width = avail_width
            return avail_width, self.height

        def draw(self) -> None:
            # Capture absolute page position of this cell's bottom-left corner
            pos = self.canv.absolutePosition(0.0, 0.0)  # type: ignore[attr-defined]
            out.append((
                int(self.canv.getPageNumber()),  # type: ignore[attr-defined]
                float(pos[0]),  # type: ignore[index]
                float(pos[1]),  # type: ignore[index]
                float(self.width),
                float(self.height),
            ))
            # Draw the label at the bottom of the cell
            p = Paragraph(label, label_style)  # type: ignore[arg-type]
            p.wrap(self.width, self.height)
            p.drawOn(self.canv, 0, 2)  # type: ignore[arg-type]  # 2 pt from bottom

    return _SigCapture()



def _add_sig_widget(
    pdf_bytes: bytes,
    name: str,
    page_number: int,
    rect: tuple[float, float, float, float],
) -> bytes:
    """
    Post-process an existing PDF (bytes) and embed an interactive AcroForm
    digital-signature field (/Sig) at the given rectangle on the given page.

    Args:
        pdf_bytes:   Source PDF as raw bytes.
        name:        Internal field name (no spaces).
        page_number: 1-indexed page number.
        rect:        (x1, y1, x2, y2) in PDF page coordinates (origin bottom-left).

    Returns:
        Modified PDF bytes with the /Sig widget annotation added.
    """
    from pypdf import PdfWriter
    from pypdf.generic import (
        ArrayObject,
        DictionaryObject,
        NameObject,
        NumberObject,
        TextStringObject,
    )

    writer = PdfWriter(clone_from=io.BytesIO(pdf_bytes))
    page = writer.pages[page_number - 1]

    # Build the widget annotation dictionary for a /Sig field
    sig_widget = DictionaryObject({
        NameObject("/Type"): NameObject("/Annot"),
        NameObject("/Subtype"): NameObject("/Widget"),
        NameObject("/FT"): NameObject("/Sig"),
        NameObject("/T"): TextStringObject(name),
        NameObject("/Rect"): ArrayObject([
            NumberObject(round(rect[0], 3)),
            NumberObject(round(rect[1], 3)),
            NumberObject(round(rect[2], 3)),
            NumberObject(round(rect[3], 3)),
        ]),
        NameObject("/F"): NumberObject(4),   # print flag
        NameObject("/DR"): DictionaryObject(),
    })
    sig_ref = writer._add_object(sig_widget)  # type: ignore[attr-defined]

    # Link back-reference from annotation to page
    sig_widget[NameObject("/P")] = page.indirect_reference  # type: ignore[assignment]

    # Add widget to page /Annots array
    if "/Annots" not in page:
        page[NameObject("/Annots")] = ArrayObject()
    page[NameObject("/Annots")].append(sig_ref)  # type: ignore[union-attr]

    # Ensure the document's AcroForm exists and register the field
    catalog = writer._root_object  # type: ignore[attr-defined]
    if NameObject("/AcroForm") not in catalog:
        catalog[NameObject("/AcroForm")] = DictionaryObject()
    acroform: DictionaryObject = catalog[NameObject("/AcroForm")]  # type: ignore[assignment]
    if NameObject("/Fields") not in acroform:
        acroform[NameObject("/Fields")] = ArrayObject()
    acroform[NameObject("/Fields")].append(sig_ref)  # type: ignore[union-attr]
    # SigFlags=3: AppendOnly (bit 2) + SignaturesExist (bit 1) — tells readers this doc has sig fields
    acroform[NameObject("/SigFlags")] = NumberObject(3)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def generate_pdf(
    entries: list[TimesheetEntryRead],
    employee_name: str,
    from_date: date,
    to_date: date,
) -> bytes:
    """
    Generate a PDF report from the given timesheet entries.

    Uses reportlab to produce an A4 PDF with:
      - Company/app title and report metadata at the top
      - A styled table with all entries (alternating row colours)
      - A totals row at the bottom of the table
      - Page numbers in the footer

    Args:
        entries:       List of timesheet entries to export.
        employee_name: Display name of the employee.
        from_date:     Start of the export date range.
        to_date:       End of the export date range.

    Returns:
        PDF file content as bytes.
    """
    # Import here to avoid loading reportlab unless a PDF export is requested
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Flowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()

    # --- Document setup ---
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    elements: list[object] = []

    # Style for description cells — enables word wrapping at font size 9
    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontSize=9,
        leading=11,
    )

    # --- Title ---
    elements.append(Paragraph("Timesheet Export", styles["Title"]))
    elements.append(Spacer(1, 0.3 * cm))

    # --- Metadata ---
    elements.append(Paragraph(f"<b>Employee:</b> {employee_name}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Period:</b> {from_date} - {to_date}", styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    # --- Table ---
    # Header row — plain strings are fine here (no wrapping needed)
    table_data: list[list[str | Paragraph]] = [
        ["Date", "Duration (hh:mm)", "Description"],
    ]

    total_minutes = 0
    for entry in entries:
        table_data.append([
            entry.entry_date.isoformat(),
            entry.hours_display,
            Paragraph(entry.description, cell_style),  # Paragraph enables word wrap
        ])
        total_minutes += entry.minutes

    # Totals row
    table_data.append([
        "TOTAL",
        _minutes_to_display(total_minutes),
        "",
    ])

    # Available page width
    page_width = A4[0] - 4 * cm
    col_widths = [3 * cm, 3.5 * cm, page_width - 6.5 * cm]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table styling
    row_count = len(table_data)
    # reportlab style commands are 4- or 5-tuples (the GRID command includes a line-width arg)
    table_style_commands: list[tuple[str, object, object, object] | tuple[str, object, object, object, object]] = [
        # Header row style
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4F8A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        # Data rows
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        # Alternating row colours
        *[
            ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#DCE6F1"))
            for row_idx in range(2, row_count - 1, 2)
        ],
        # Totals row — last data row before the end
        ("BACKGROUND", (0, row_count - 1), (-1, row_count - 1), colors.HexColor("#BDD7EE")),
        ("FONTNAME", (0, row_count - 1), (-1, row_count - 1), "Helvetica-Bold"),
    ]

    table.setStyle(TableStyle(table_style_commands))  # type: ignore[arg-type]
    elements.append(table)

    # --- Signature section ---
    # Gap + two-column layout: Datum (left) | gap | Unterschrift Kunde (right, digital sig)
    elements.append(Spacer(1, 1.5 * cm))

    sig_label_style = ParagraphStyle(
        "SigLabel",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#555555"),
    )

    sig_height = 1.2 * cm          # height of the signature row
    gap_width = 1.0 * cm           # space between the two fields
    page_width = A4[0] - 4 * cm    # available width (same as table)
    sig_col_width = (page_width - gap_width) / 2

    # Capture list — the _SigCapture flowable appends its abs position here during draw()
    sig_positions: list[tuple[int, float, float, float, float]] = []

    sig_data: list[list[object]] = [[
        Paragraph("Datum", sig_label_style),
        "",   # gap column — no border, no content
        _make_sig_capture_flowable(Flowable, "Unterschrift Kunde", sig_height, sig_label_style, sig_positions),
    ]]
    sig_table = Table(
        sig_data,
        colWidths=[sig_col_width, gap_width, sig_col_width],
        rowHeights=[sig_height],
    )
    sig_table.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (0, 0), 0.75, colors.black),   # line above "Datum"
        ("LINEABOVE", (2, 0), (2, 0), 0.75, colors.black),   # line above "Unterschrift Kunde"
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
    ]))
    elements.append(sig_table)

    # --- Build ---
    doc.build(elements)  # type: ignore[arg-type]
    pdf_bytes = buffer.getvalue()

    # --- Post-process: add interactive /Sig AcroForm field ---
    if sig_positions:
        page_no, abs_x, abs_y, field_w, field_h = sig_positions[0]
        pdf_bytes = _add_sig_widget(
            pdf_bytes,
            name="Unterschrift_Kunde",
            page_number=page_no,
            rect=(abs_x, abs_y, abs_x + field_w, abs_y + field_h),
        )

    return pdf_bytes
